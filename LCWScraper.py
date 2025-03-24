from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException, TimeoutException
import time
import os
import re
from datetime import datetime
from openpyxl import Workbook #Excel işlemleri için kullanılır
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter


class StoreInfo:
    def __init__(self, country, store_name, address, working_hours, phone): #Çekilen verileri saklamak için kullanılır
        self.country = country
        self.store_name = store_name
        self.address = address
        self.working_hours = working_hours
        self.phone = phone


class LCWaikikiStoreScraper:
    def __init__(self): #Başlangıç ayarları yapılır ve chrome açılır.
        self.driver = None
        self.wait = None
        self.js = None

    def initialize(self):
        options = Options()
        options.add_argument("--start-maximized")
        options.add_argument("--disable-gpu")

        service = Service()
        self.driver = webdriver.Chrome(service=service, options=options)
        self.wait = WebDriverWait(self.driver, 10)
        self.js = self.driver

    def scrape_stores(self):
        self.initialize()
        stores_by_country = {}

        try: #Programın ana metodu
            self.driver.get("https://corporate.lcwaikiki.com/magazalar")

            country_select = self.wait_and_get_element(By.ID, "DDLCountry")
            select_element = Select(country_select)
            countries = [option.text for option in select_element.options
                         if option.text != "Ülke" and option.text != "Seçiniz"]

            for country in countries:
                try:
                    country_stores = self.scrape_country_stores(select_element, country)
                    if country_stores:
                        stores_by_country[country] = country_stores
                except Exception as ex:
                    print(f"Ülke işlenirken hata: {country} - {str(ex)}")

        except Exception as ex:
            print(f"Genel scraping hatası: {str(ex)}")
        finally:
            if self.driver:
                self.driver.quit()

        return stores_by_country

    def scrape_country_stores(self, select_element, country):
        country_stores = []
        print(f"\n {country} ülkesindeki mağazalar yükleniyor")

        select_element.select_by_visible_text(country)
        time.sleep(1)  # Sayfanın yüklenmesi için bekleriz

        store_elements = self.driver.find_elements(By.CSS_SELECTOR, "#ListArea li")

        if not store_elements:
            print(f" {country} için mağaza bulunamadı.")
            return country_stores

        print(f" {country} için {len(store_elements)} mağaza bulundu")

        # Ana mağaza veri çekme fonksiyonu
        self.process_store_elements(store_elements, country_stores, country)

        # Scroll varsa ek mağazaları yükler
        self.try_load_more_stores_with_scroll(country_stores, country)

        print(f" {country} ülkesinde toplam {len(country_stores)} mağaza bilgisi alındı.")
        return country_stores

    def process_store_elements(self, store_elements, country_stores, country):
        for store_element in store_elements: #Bilgileri çekip ekrana yazdığımız kısım
            try:
                self.js.execute_script("arguments[0].scrollIntoView({block: 'center'});", store_element)
                time.sleep(0.2)

                store_name = self.extract_store_name(store_element)
                address, working_hours, phone = self.extract_store_details(store_element)

                country_stores.append(StoreInfo(
                    country=country,
                    store_name=store_name,
                    address=address,
                    working_hours=working_hours,
                    phone=phone
                ))

                print(f" {store_name} → Adres: {address} | Çalışma Saatleri: {working_hours} | Telefon: {phone}")
            except Exception as ex:
                print(f"Mağaza bilgisi alınırken hata: {str(ex)}")

    def try_load_more_stores_with_scroll(self, country_stores, country):
        try: #Mağaza elementlerini kontrol eder
            scroll_pane = self.driver.find_elements(By.CLASS_NAME, "jspPane")
            scroll_bar = self.driver.find_elements(By.CLASS_NAME, "jspDrag")

            if scroll_pane and scroll_bar:
                previous_store_count = len(country_stores)
                max_scroll_tries = 10

                for i in range(max_scroll_tries):
                    scroll_amount = 500  # Sabit scroll miktarı
                    self.js.execute_script(f"document.getElementById('ListArea').scrollTop += {scroll_amount};")
                    time.sleep(0.8)

                    new_store_elements = self.driver.find_elements(By.CSS_SELECTOR, "#ListArea li")
                    if len(new_store_elements) <= previous_store_count:
                        break  # Yeni mağazalar yüklenmediyse döngüyü sonlandır

                    # Yeni yüklenen mağazaları işle
                    new_elements = new_store_elements[previous_store_count:]
                    self.process_store_elements(new_elements, country_stores, country)

                    previous_store_count = len(new_store_elements)
        except Exception as ex:
            print(f"Scroll işleminde hata: {str(ex)}")

    def extract_store_name(self, store_element):
        try:
            return store_element.find_element(By.TAG_NAME, "h3").text.strip()
        except:
            return "Bulunmadı"

    def extract_store_details(self, store_element):
        address = "Bulunmadı"
        working_hours = "Bulunmadı"
        phone = "Bulunmadı"

        try:
            paragraph_elements = store_element.find_elements(By.TAG_NAME, "p")
            paragraph_texts = [p.text.strip() for p in paragraph_elements if p.text.strip()]

            # Telefon numarasını tespit et
            for text in paragraph_texts:
                if "Tel:" in text or "Telefon:" in text or re.search(r"\+?\d[\d\s-]{7,}", text):
                    phone = text
                    break

            # Çalışma saatlerini tespit et
            for text in paragraph_texts:
                if ("Pazartesi" in text or "Salı" in text or
                        "Çarşamba" in text or "Perşembe" in text or
                        "Cuma" in text or "Cumartesi" in text or
                        "Pazar" in text or
                        (":" in text and re.search(r"\d{1,2}[:.-]\d{2}", text)) or
                        re.search(r"\d{1,2}(:|\.)\d{2}\s*(-|–)\s*\d{1,2}(:|\.)\d{2}", text) or
                        "AM" in text or "PM" in text or
                        "Çalışma Saatleri" in text or "Haftaiçi" in text or
                        "Haftasonu" in text or "Açılış" in text or
                        "Kapanış" in text):
                    working_hours = text
                    break

            # Adresi tespit et
            for text in paragraph_texts:
                if text != phone and text != working_hours:
                    address = text
                    break

        except Exception as ex:
            print(f"Paragraf işleme hatası: {str(ex)}")

        return address, working_hours, phone

    def wait_and_get_element(self, by, selector):
        #Istenen elementlerin sayfada görünmesini bekler
        try:
            return self.wait.until(EC.presence_of_element_located((by, selector)))
        except TimeoutException:
            print(f"Element bulunamadı: {selector}")
            return None

    def export_to_excel(self, stores_by_country):
        #Excel'e aktarma kısmı
        desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
        file_path = os.path.join(desktop_path, f"Magazalar_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Mağaza Bilgileri"


        headers = ["Ülke", "Mağaza Adı", "Adres", "Çalışma Saatleri", "Telefon"]
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        current_row = 2

        # Veri girişi
        for country in sorted(stores_by_country.keys()):
            stores = stores_by_country[country]
            if not stores:
                continue

            worksheet.cell(row=current_row, column=1).value = country
            worksheet.cell(row=current_row, column=2).value = stores[0].store_name
            worksheet.cell(row=current_row, column=3).value = stores[0].address
            worksheet.cell(row=current_row, column=4).value = stores[0].working_hours
            worksheet.cell(row=current_row, column=5).value = stores[0].phone

            for i in range(1, len(stores)):
                current_row += 1
                worksheet.cell(row=current_row, column=2).value = stores[i].store_name
                worksheet.cell(row=current_row, column=3).value = stores[i].address
                worksheet.cell(row=current_row, column=4).value = stores[i].working_hours
                worksheet.cell(row=current_row, column=5).value = stores[i].phone

            # Ülke hücrelerini birleştirir
            if len(stores) > 1:
                worksheet.merge_cells(start_row=current_row - (len(stores) - 1), start_column=1, end_row=current_row,
                                      end_column=1)
                worksheet.cell(row=current_row - (len(stores) - 1), column=1).alignment = Alignment(vertical="center")

            # Gruplar arasında boşluk bırakır
            current_row += 1

        # Sütun genişliklerini ayarlar
        column_widths = [20, 30, 40, 30, 25]
        for i, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[get_column_letter(i)].width = width

        # Hücrelerin metni sarması için ayar yaparız
        for row_cells in worksheet.iter_rows():  # Değişken adını değiştirdim
            for cell in row_cells:
                cell.alignment = Alignment(wrap_text=True)

        # Tüm hücrelere kenarlık ekleriz
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        max_row = current_row - 1  # Son satır numarasını belirler
        for row_cells in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1,
                                             max_col=5):  # Değişken adını değiştirdim
            for cell in row_cells:
                cell.border = thin_border

        workbook.save(file_path)
        print(f"Veriler {file_path} konumuna kaydedildi.")

    @staticmethod
    def main():
        scraper = LCWaikikiStoreScraper()
        stores_by_country = scraper.scrape_stores()

        # Ülke bazında istatistikler
        print("\n===== Ülkelere Göre Mağaza Sayıları =====")
        for country in sorted(stores_by_country.keys()):
            print(f"{country}: {len(stores_by_country[country])} mağaza")

        total_stores = sum(len(stores) for stores in stores_by_country.values())
        total_countries = len(stores_by_country)
        print(f"\nToplam {total_countries} ülkede {total_stores} mağaza bilgisi alındı.")

        scraper.export_to_excel(stores_by_country)


if __name__ == "__main__":
    LCWaikikiStoreScraper.main()